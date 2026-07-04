# -*- coding: utf-8 -*-
"""
MDC AI月次着地予測システム — クラウド閲覧専用版（cloud_viewer）
====================================================================
これは院内検証用のクラウド閲覧専用画面です。
- 予測更新ボタンなし / run_all.bat 実行なし / 管理者モードなし / ログ表示なし
- 表示するのは、data/ に同梱した「集計済みoutputs」だけ（個票・患者情報は一切含まない）
- 正ダッシュボード本体＝この画面。dashboard_v3 は共有・保存用の出力レポート（メインにしない）
- デザインは MDC Forecast Console（紺×ゴールド・大型カード・視線誘導）を踏襲

起動: py -m streamlit run streamlit_app.py
"""
import os
import html as _html
import streamlit as st

# ----------------------------------------------------------------------
# 基本設定（ローカルパスは一切出さない。data/ 相対のみ参照）
# ----------------------------------------------------------------------
BASE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(BASE, "data")
DATA_DATE = "2026年7月4日"          # データ作成日（レポート作成日時が不明な場合の表示）
TARGET_YM_JP = "2026年7月"
FALLBACK_PW = "mdc202607"

# dashboard_v3_2026_07.xlsx のシート名（openpyxlで正しく読める日本語名）
XLSX = os.path.join(DATA, "dashboard_v3_2026_07.xlsx")
SHEET_COMPARE = "前年比較"

# forecast_model_v2 はクラウド版に同梱しない（個票由来のため）。
# 80%レンジ・高単価型レンジは、同梱summary/レポートに記載の確定表示値を定数で持つ（推測ではない）。
RANGE80 = (17120000, 20580000)      # 80%予測レンジ 1,712〜2,058万円
HIGH_VALUE_RANGE = (2500000, 3840000)   # 高単価型自費レンジ 250〜384万円

st.set_page_config(page_title="MDC AI月次着地予測システム", page_icon="📈", layout="wide")

# ======================================================================
# 表示ヘルパー
# ======================================================================
def man(v):
    try:
        return f"{round(float(v) / 10000):,}万円"
    except Exception:
        return "取得不可"

def manv(v):
    try:
        return f"{round(float(v) / 10000):,}"
    except Exception:
        return "—"

def sman(v):
    try:
        n = round(float(v) / 10000)
        return (f"▲{abs(n):,}万円" if n < 0 else (f"+{n:,}万円" if n > 0 else "±0万円"))
    except Exception:
        return "取得不可"

def signclass(v):
    try:
        n = float(v)
        return "dn" if n < 0 else ("up" if n > 0 else "fl")
    except Exception:
        return "fl"

def judgeclass(txt):
    t = txt or ""
    if "超え" in t:
        return "up"
    if "割れ" in t or "高い" in t:
        return "dn"
    return "fl"

# ======================================================================
# パスワード保護
#   優先度: st.secrets["VIEW_PASSWORD"] > 環境変数 VIEW_PASSWORD > 仮パス mdc202607
# ======================================================================
def expected_password():
    try:
        if "VIEW_PASSWORD" in st.secrets:
            return str(st.secrets["VIEW_PASSWORD"])
    except Exception:
        pass
    env = os.environ.get("VIEW_PASSWORD")
    if env:
        return env
    return FALLBACK_PW

def check_password():
    if st.session_state.get("_authed"):
        return True
    st.markdown(
        "<div style='max-width:420px;margin:8vh auto 0;text-align:center;'>"
        "<div style='font-size:22px;font-weight:800;color:#0B1F3A;'>MDC AI月次着地予測システム</div>"
        "<div style='font-size:13px;color:#6b7686;margin:8px 0 18px;'>院内検証用・クラウド閲覧専用画面</div>"
        "</div>",
        unsafe_allow_html=True,
    )
    c = st.columns([1, 2, 1])[1]
    with c:
        pw = st.text_input("閲覧パスワード", type="password", key="_pw_input")
        if st.button("表示する", type="primary", width="stretch"):
            if pw == expected_password():
                st.session_state["_authed"] = True
                st.rerun()
            else:
                st.error("パスワードが違います。")
        st.caption("パスワードは院長・事務局にご確認ください。")
    return False

# ======================================================================
# データ読み込み（同梱xlsxの「前年比較」シートのみ）
# ======================================================================
@st.cache_data(show_spinner=False)
def read_metrics():
    """dashboard_v3_2026_07.xlsx の 前年比較 シートを {項目: {...}} で返す。"""
    m = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(XLSX, data_only=True)
        ws = wb[SHEET_COMPARE]
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0] is not None:
                m[str(r[0])] = dict(pred=r[1], py=r[2], diff=r[3], rate=r[4], judge=r[5])
        wb.close()
    except Exception:
        return None
    return m

@st.cache_data(show_spinner=False)
def read_text(fname):
    p = os.path.join(DATA, fname)
    try:
        with open(p, encoding="utf-8", errors="replace") as f:
            return f.read()
    except Exception:
        return None

# ======================================================================
# CSS（dashboard_v3 トンマナ：紺×ゴールド・落ち着いた赤）+ スマホ対応
# ======================================================================
CSS = """
<style>
:root{--navy:#0B1F3A;--navy2:#1A3358;--gold:#C8A96A;--red:#B5544A;--ink:#1E1E1E;--muted:#6b7686;}
[data-testid="stDecoration"]{display:none;}
[data-testid="stHeader"]{background:rgba(255,255,255,0);height:auto;}
.block-container{padding-top:3.2rem !important;padding-bottom:2rem;max-width:1360px;overflow:visible;}
.mfc-title{font-size:30px;font-weight:800;color:var(--navy);letter-spacing:.3px;line-height:1.3;margin:.3rem 0 6px;}
.mfc-sub{font-size:13px;color:var(--muted);margin-bottom:6px;line-height:1.6;}
.mfc-meta{font-size:12px;color:var(--muted);margin-bottom:14px;}
.mfc-meta b{color:var(--navy);}
.mfc-hero{background:linear-gradient(135deg,#0B1F3A 0%,#1A3358 100%);border-radius:16px;padding:20px 24px;color:#fff;box-shadow:0 5px 16px rgba(11,31,58,.28);}
.mfc-concl{font-size:16px;color:#e6ecf5;line-height:1.7;margin:2px 0 16px;}
.mfc-concl b{color:#fff;}.mfc-concl .g{color:var(--gold);font-weight:800;}
.mfc-grid6{display:grid;grid-template-columns:1.5fr 1fr 1fr 1.2fr 1.2fr 1.4fr;gap:11px;}
.mfc-hc{background:rgba(255,255,255,.05);border:1px solid #33507a;border-radius:12px;padding:11px 13px;}
.mfc-hc .lb{font-size:11px;color:var(--gold);font-weight:700;margin-bottom:5px;letter-spacing:.2px;}
.mfc-hc .vl{font-size:25px;font-weight:800;color:#fff;line-height:1.05;}
.mfc-hc .vl .u{font-size:12px;color:#aeb9c9;margin-left:2px;font-weight:700;}
.mfc-hc.main{background:rgba(200,169,106,.15);border-color:var(--gold);}
.mfc-hc.main .vl{font-size:33px;}
.mfc-up{color:#7FE0A6 !important;}.mfc-dn{color:#FF9E9E !important;}.mfc-fl{color:#F0C674 !important;}
.mfc-judge{border-left:5px solid var(--gold);background:#f6f7f9;border-radius:0 10px 10px 0;padding:12px 16px;margin:16px 0 4px;font-size:14px;color:#333;line-height:1.7;}
.mfc-judge b{color:var(--navy);}
.mfc-judge ul{margin:6px 0 0;padding-left:20px;}
.mfc-judge li{margin:3px 0;}
.mfc-badge{display:inline-block;font-size:12px;font-weight:800;padding:2px 10px;border-radius:9px;margin-left:6px;}
.b-up{background:#e7f5ec;color:#2E7D57;}.b-dn{background:#f6e9e7;color:#B5544A;}.b-fl{background:#fbf1df;color:#9a7420;}
.mfc-sec{font-size:16px;font-weight:800;color:var(--navy);border-left:6px solid var(--gold);padding-left:10px;margin:24px 0 10px;}
.mfc-cards{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;}
.mfc-card{background:#fff;border:1px solid #e3e7ee;border-radius:14px;padding:15px 17px;box-shadow:0 1px 5px rgba(0,0,0,.05);}
.mfc-card .lb{font-size:12px;font-weight:700;color:var(--navy);margin-bottom:6px;}
.mfc-card .big{font-size:29px;font-weight:800;color:var(--navy);line-height:1;}
.mfc-card .big .u{font-size:13px;color:#8a94a3;margin-left:3px;}
.mfc-card .sub{font-size:11px;color:var(--muted);margin-top:6px;}
.mfc-hv{grid-column:span 3;background:linear-gradient(135deg,#fbf4e6,#fff);border:1px solid var(--gold);border-radius:14px;padding:15px 18px;display:flex;align-items:center;gap:18px;box-shadow:0 1px 5px rgba(0,0,0,.05);}
.mfc-hv .lb{font-size:13px;font-weight:800;color:#7a5a10;}
.mfc-hv .rng{font-size:30px;font-weight:800;color:var(--navy);}
.mfc-hv .rng .u{font-size:13px;color:#8a94a3;margin-left:3px;}
.mfc-hv .note{font-size:12px;color:#7a5a10;flex:1;}
.mfc-tw{width:100%;overflow-x:auto;}
table.mfc-t{width:100%;border-collapse:collapse;font-size:13.5px;min-width:520px;}
table.mfc-t th{background:var(--navy);color:#fff;padding:8px 10px;text-align:right;font-weight:700;}
table.mfc-t th:first-child{text-align:left;}
table.mfc-t td{border-bottom:1px solid #eef0f4;padding:8px 10px;text-align:right;color:#333;}
table.mfc-t td:first-child{text-align:left;font-weight:700;color:var(--navy);}
table.mfc-t tr:hover td{background:#fafbfc;}
.t-dn{color:#B5544A;font-weight:700;}.t-up{color:#2E7D57;font-weight:700;}.t-fl{color:#9a7420;font-weight:700;}
.mfc-actions{background:#fff;border:1px solid #e3e7ee;border-top:4px solid var(--gold);border-radius:14px;padding:16px 20px;box-shadow:0 1px 6px rgba(0,0,0,.06);}
.mfc-actions .h{font-size:15px;font-weight:800;color:var(--navy);margin-bottom:10px;}
.mfc-actions ul{list-style:none;margin:0;padding:0;}
.mfc-actions li{font-size:14px;color:#2b2b2b;padding:9px 0 9px 30px;position:relative;border-bottom:1px dashed #ececf0;line-height:1.5;}
.mfc-actions li:last-child{border-bottom:none;}
.mfc-actions li:before{content:'▢';position:absolute;left:4px;color:var(--gold);font-weight:800;}
.mfc-actions li b{color:var(--navy);}
.mfc-note{font-size:12px;color:var(--muted);margin-top:8px;line-height:1.6;}
/* --- スマホ幅：カード・ヒーローを縦積み --- */
@media (max-width:820px){
  .mfc-grid6{grid-template-columns:1fr 1fr;}
  .mfc-hc.main{grid-column:span 2;}
  .mfc-cards{grid-template-columns:1fr;}
  .mfc-hv{grid-column:span 1;flex-direction:column;align-items:flex-start;gap:8px;}
  .mfc-title{font-size:24px;}
  .mfc-hc .vl{font-size:22px;}.mfc-hc.main .vl{font-size:28px;}
}
</style>
"""

# ======================================================================
# 本体描画
# ======================================================================
def render():
    st.markdown(CSS, unsafe_allow_html=True)
    M = read_metrics()

    # ---------- タイトル・説明・更新日 ----------
    st.markdown('<div class="mfc-title">MDC AI月次着地予測システム</div>', unsafe_allow_html=True)
    st.markdown(
        "<div class='mfc-sub'>これは院内検証用のクラウド閲覧専用画面です。"
        "表示値は確定値ではなく、経営判断のための推定値です。</div>",
        unsafe_allow_html=True)
    st.markdown(
        f"<div class='mfc-meta'>データ作成日：<b>{DATA_DATE}</b>　｜　対象月：<b>{TARGET_YM_JP}</b>　"
        "｜　正ダッシュボード本体＝この画面（dashboard_v3 は共有・保存用の出力レポート）</div>",
        unsafe_allow_html=True)

    if not M:
        st.warning("表示用データ（dashboard_v3_2026_07.xlsx）を読み込めませんでした。")
        return

    cur = M.get("月間総売上(現時点着地見込み)", {})
    base = M.get("月間総売上(通常営業ベース)", {})
    try:
        diff_base = float(cur.get("pred")) - float(base.get("pred"))
    except Exception:
        diff_base = None
    py_diff = cur.get("diff")

    # ---------- 4. 結論1行 + 5. ヒーローカード ----------
    concl = (f"<b>{TARGET_YM_JP}</b> の現時点着地見込みは <span class='g'>{man(cur.get('pred'))}</span>。"
             f"前年同月比では <span class='mfc-{signclass(py_diff)}'>{sman(py_diff)}</span>、"
             f"通常営業ベースとの差は <span class='mfc-{signclass(diff_base)}'>{sman(diff_base)}</span> です。")

    def hc(lb, num, unit="万円", cls="", numcls=""):
        return (f"<div class='mfc-hc {cls}'><div class='lb'>{lb}</div>"
                f"<div class='vl {numcls}'>{num}<span class='u'>{unit}</span></div></div>")

    hero = (
        "<div class='mfc-hero'>"
        f"<div class='mfc-concl'>{concl}</div>"
        "<div class='mfc-grid6'>"
        + hc("現時点着地見込み", manv(cur.get("pred")), "万円", "main")
        + hc("前年同月実績", manv(cur.get("py")))
        + hc("前年差", sman(py_diff).replace("万円", ""), "万円", "", f"mfc-{signclass(py_diff)}")
        + hc("通常営業ベース予測", manv(base.get("pred")))
        + hc("通常営業ベースとの差", sman(diff_base).replace("万円", ""), "万円", "", f"mfc-{signclass(diff_base)}")
        + hc("80%予測レンジ", f"{manv(RANGE80[0])}〜{manv(RANGE80[1])}")
        + "</div></div>"
    )
    st.markdown(hero, unsafe_allow_html=True)

    # ---------- 6. 判定エリア ----------
    st.markdown(
        "<div class='mfc-judge'>現時点では<b>前年同月を下回る見込み</b>です。"
        f"<span class='mfc-badge b-{judgeclass(cur.get('judge'))}'>{cur.get('judge') or '—'}</span>"
        "<ul>"
        "<li>ただし通常営業ベースとの差は、月末実績で再検証します。</li>"
        "<li>木曜休診の影響は、あくまで候補として見ます。</li>"
        "<li>確定的な損失とは扱いません（月末後に吸収されたかを判定）。</li>"
        "</ul></div>",
        unsafe_allow_html=True)

    # ---------- 7. 売上構成 ----------
    st.markdown('<div class="mfc-sec">売上構成（当月予測）</div>', unsafe_allow_html=True)

    def scard(lb, key, sub=""):
        v = M.get(key, {}).get("pred")
        return (f"<div class='mfc-card'><div class='lb'>{lb}</div>"
                f"<div class='big'>{manv(v)}<span class='u'>万円</span></div>"
                f"<div class='sub'>{sub}</div></div>")

    hv_html = (
        f"<div class='mfc-hv'><div><div class='lb'>高単価型 自費レンジ</div>"
        f"<div class='rng'>{manv(HIGH_VALUE_RANGE[0])}〜{manv(HIGH_VALUE_RANGE[1])}<span class='u'>万円</span></div></div>"
        "<div class='note'>高単価型自費は月末着地に与える影響が大きいため、"
        "案件別の進捗確認が必要です（上振れ要因）。</div></div>"
    )
    st.markdown(
        "<div class='mfc-cards'>"
        + scard("保険診療売上予測", "保険診療売上", "安定指標（基礎売上の芯）")
        + scard("自費診療売上予測", "自費診療売上", "変動が大きく差の主因になりうる")
        + scard("物販売上予測", "物販売上", "影響は小さい")
        + hv_html + "</div>", unsafe_allow_html=True)

    # ---------- 8. 前年比較表 ----------
    st.markdown('<div class="mfc-sec">前年比較</div>', unsafe_allow_html=True)

    def cell_cnt(v, u):
        try:
            return f"{round(float(v)):,}{u}"
        except Exception:
            return "取得不可"

    def build_rows():
        rows = []
        specs = [("総売上（着地見込み）", "月間総売上(現時点着地見込み)", "money", ""),
                 ("保険", "保険診療売上", "money", ""), ("自費", "自費診療売上", "money", ""),
                 ("物販", "物販売上", "money", ""), ("来院", "総来院回数", "cnt", "回"),
                 ("初診", "初診件数", "cnt", "件")]
        for label, key, kind, u in specs:
            r = M.get(key)
            if not r:
                rows.append(f"<tr><td>{label}</td><td>取得不可</td><td>取得不可</td>"
                            "<td>取得不可</td><td>取得不可</td><td>取得不可</td></tr>")
                continue
            if kind == "money":
                pred, py = man(r.get("pred")), man(r.get("py"))
                dv = sman(r.get("diff"))
            else:
                pred, py = cell_cnt(r.get("pred"), u), cell_cnt(r.get("py"), u)
                try:
                    n = round(float(r.get("diff")))
                    dv = (f"▲{abs(n):,}" if n < 0 else (f"+{n:,}" if n > 0 else "±0"))
                except Exception:
                    dv = "取得不可"
            dcls = "t-" + signclass(r.get("diff"))
            rt = f"{r.get('rate')}%" if r.get("rate") is not None else "取得不可"
            jb = f"<span class='mfc-badge b-{judgeclass(r.get('judge'))}'>{r.get('judge') or '—'}</span>"
            rows.append(f"<tr><td>{label}</td><td>{pred}</td><td>{py}</td>"
                        f"<td class='{dcls}'>{dv}</td><td class='{dcls}'>{rt}</td>"
                        f"<td style='text-align:right'>{jb}</td></tr>")
        return "".join(rows)

    st.markdown(
        "<div class='mfc-tw'><table class='mfc-t'><thead><tr><th>項目</th><th>今月予測</th><th>前年同月</th>"
        "<th>差分</th><th>差率</th><th>判定</th></tr></thead><tbody>"
        + build_rows() + "</tbody></table></div>"
        "<div class='mfc-note'>前年同月実績（確定）との比較。読み取れない項目は「取得不可」と表示し、推測はしません。</div>",
        unsafe_allow_html=True)

    # ---------- 9. 今月、月末までに確認すること ----------
    st.markdown('<div class="mfc-sec">今月、月末までに確認すること</div>', unsafe_allow_html=True)
    acts = [
        "高単価型自費案件の進捗確認",
        "月内売上化できる案件と翌月送りになる案件の仕分け",
        "継続管理型の未充足枠・キャンセル枠の補充",
        "初診から自費相談・治療移行につながる案件の確認",
        "月末確定後、通常営業ベースとの差が吸収されたか判定",
    ]
    li = "".join(f"<li><b>{i}.</b> {_html.escape(a)}</li>" for i, a in enumerate(acts, 1))
    st.markdown(
        "<div class='mfc-actions'><div class='h'>院長・事務局が「で、何をするか」を見る場所です</div>"
        f"<ul>{li}</ul></div>", unsafe_allow_html=True)

    # ---------- 10. 出力レポート確認（折りたたみ・メインにしない） ----------
    with st.expander("出力レポート確認（共有・保存用）", expanded=False):
        st.caption("以下は共有・保存用に自動生成されたレポートです。正ダッシュボード本体はこの画面です。")
        png = os.path.join(DATA, "dashboard_v3_2026_07.png")
        if os.path.exists(png):
            st.image(png, width="stretch",
                     caption="dashboard_v3（出力レポート／共有・保存用）")
        summary = read_text("dashboard_v3_summary_2026_07.md")
        if summary:
            with st.expander("summary.md の内容", expanded=False):
                st.markdown(summary)
        fs = read_text("forecast_summary_v2_2026_07.md")
        if fs:
            with st.expander("予測根拠サマリー（forecast_summary_v2）", expanded=False):
                st.markdown(fs)
        mc = read_text("model_card_v2_2026_07.md")
        if mc:
            with st.expander("モデル説明資料（model_card_v2）", expanded=False):
                st.markdown(mc)

    # ---------- 11. 注意書き ----------
    with st.expander("注意・限界（必ずお読みください）", expanded=False):
        st.markdown(
            "- 表示値は確定値ではなく、経営判断のための推定値です。\n"
            "- 月中予測は暫定です（月初予測ロジックで実行）。\n"
            "- 自費は変動が大きく、差の主因になりえます。\n"
            "- 高単価型は案件別の確認が必要です。\n"
            "- 通常営業ベースとの差は、月末後に実績と比較して再検証します（確定的な損失ではありません）。\n"
            "- 本画面は院内検証用です。")

    st.divider()
    st.caption("MDC AI月次着地予測システム（クラウド閲覧専用）｜院内検証用｜"
               "表示値は推定値・確定値ではありません｜個人情報・患者番号は非表示")


# ======================================================================
# エントリポイント：パスワード通過後のみ本体を描画
# ======================================================================
if check_password():
    render()
